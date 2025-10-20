"""
Excel模板生成器
用于生成各种标准Excel模板
"""

import os
import tempfile
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger


class MaterialTemplateGenerator:
    """材料模板生成器"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def generate_base_material_template(self) -> str:
        """生成基准材料导入模板"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "基准材料模板"
            
            # 设置标题行
            headers = [
                "材料名称*", "规格型号", "单位*", "当前价格*", "材料分类*", 
                "适用地区", "数据来源", "生效日期", "备注", "是否常用"
            ]
            
            # 添加标题说明
            ws.merge_cells('A1:J1')
            title_cell = ws['A1']
            title_cell.value = "造价材料审计系统 - 基准材料导入模板"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加说明信息
            ws.merge_cells('A2:J2')
            desc_cell = ws['A2']
            desc_cell.value = "请按照模板格式填写材料信息，标*为必填项"
            desc_cell.font = Font(size=10, color="666666")
            desc_cell.alignment = Alignment(horizontal="center")
            
            # 设置列标题
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加示例数据
            sample_data = [
                ["商品混凝土C30", "泵送", "m³", "320.00", "混凝土", "北京", "市场价", "2025-01-01", "常用材料", "是"],
                ["钢筋HRB400", "φ12", "kg", "4.80", "钢材", "上海", "信息价", "2025-01-01", "", "是"],
                ["水泥P.O42.5", "袋装", "t", "380.00", "水泥", "广州", "厂家价", "2025-01-01", "", "否"],
                ["红砖", "240×115×53", "块", "0.35", "砖类", "深圳", "市场价", "2025-01-01", "标准砖", "是"],
                ["电线BV2.5", "铜芯", "m", "3.20", "电料", "成都", "批发价", "2025-01-01", "", "否"]
            ]
            
            for row_num, row_data in enumerate(sample_data, 4):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num).value = value
            
            # 设置列宽
            column_widths = [20, 15, 8, 12, 12, 10, 10, 12, 20, 10]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为数据区域设置边框
            for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            # 添加数据验证
            # 材料分类下拉列表
            dv_category = DataValidation(
                type="list",
                formula1='"混凝土,钢材,水泥,砖类,电料,管材,石材,木材,防水材料,保温材料,其他"'
            )
            ws.add_data_validation(dv_category)
            dv_category.add(f"E4:E1000")
            
            # 是否常用下拉列表
            dv_common = DataValidation(
                type="list",
                formula1='"是,否"'
            )
            ws.add_data_validation(dv_common)
            dv_common.add(f"J4:J1000")
            
            # 添加说明工作表
            ws_info = wb.create_sheet("填写说明")
            info_data = [
                ["字段名称", "说明", "示例"],
                ["材料名称*", "材料的标准名称，必填", "商品混凝土C30"],
                ["规格型号", "材料的具体规格", "泵送、φ12、240×115×53"],
                ["单位*", "计量单位，必填", "m³、kg、t、块、m"],
                ["当前价格*", "材料当前价格，数值格式，必填", "320.00"],
                ["材料分类*", "从下拉列表中选择，必填", "混凝土、钢材、水泥等"],
                ["适用地区", "价格适用的地区", "北京、上海、广州"],
                ["数据来源", "价格数据来源", "市场价、信息价、厂家价"],
                ["生效日期", "价格生效日期，格式：YYYY-MM-DD", "2025-01-01"],
                ["备注", "其他说明信息", "常用材料、特殊要求等"],
                ["是否常用", "是否为常用材料，从下拉列表选择", "是或否"]
            ]
            
            for row_num, row_data in enumerate(info_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_info.cell(row=row_num, column=col_num)
                    cell.value = value
                    if row_num == 1:  # 标题行
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    cell.border = thin_border
            
            # 设置说明工作表列宽
            ws_info.column_dimensions['A'].width = 15
            ws_info.column_dimensions['B'].width = 40
            ws_info.column_dimensions['C'].width = 25
            
            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"base_material_template_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"基准材料模板生成成功: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"生成基准材料模板失败: {e}")
            raise e
    
    def generate_project_material_template(self) -> str:
        """生成项目材料导入模板"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "项目材料模板"
            
            # 设置标题行
            headers = [
                "序号", "材料名称*", "规格型号", "单位*", "数量*", "单价", 
                "合价", "品牌", "材料分类", "备注"
            ]
            
            # 添加标题说明
            ws.merge_cells('A1:J1')
            title_cell = ws['A1']
            title_cell.value = "造价材料审计系统 - 项目材料清单导入模板"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加说明信息
            ws.merge_cells('A2:J2')
            desc_cell = ws['A2']
            desc_cell.value = "请按照模板格式填写项目材料清单，标*为必填项"
            desc_cell.font = Font(size=10, color="666666")
            desc_cell.alignment = Alignment(horizontal="center")
            
            # 设置列标题
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加示例数据
            sample_data = [
                [1, "商品混凝土C30", "泵送", "m³", "120.5", "320.00", "38560.00", "某某牌", "混凝土", "基础用"],
                [2, "钢筋HRB400", "φ12", "t", "5.2", "4800.00", "24960.00", "首钢", "钢材", "主筋"],
                [3, "红砖", "240×115×53", "千块", "15", "350.00", "5250.00", "本地砖厂", "砖类", "填充墙"],
                [4, "水泥P.O42.5", "袋装", "t", "8.5", "380.00", "3230.00", "海螺", "水泥", ""],
                [5, "电线BV2.5", "铜芯", "m", "500", "3.20", "1600.00", "远东", "电料", "照明回路"]
            ]
            
            for row_num, row_data in enumerate(sample_data, 4):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num).value = value
            
            # 设置列宽
            column_widths = [8, 20, 15, 8, 10, 12, 12, 12, 12, 20]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为数据区域设置边框
            for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            # 添加数据验证
            # 材料分类下拉列表
            dv_category = DataValidation(
                type="list", 
                formula1='"混凝土,钢材,水泥,砖类,电料,管材,石材,木材,防水材料,保温材料,其他"'
            )
            ws.add_data_validation(dv_category)
            dv_category.add(f"I4:I1000")
            
            # 保存文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"project_material_template_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)
            
            wb.save(filepath)
            
            logger.info(f"项目材料模板生成成功: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"生成项目材料模板失败: {e}")
            raise e


class ReportTemplateGenerator:
    """报告模板生成器"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def generate_audit_report_template(self) -> str:
        """生成审计报告模板"""
        # 这里可以实现审计报告模板生成逻辑
        pass


# 辅助函数
def clean_temp_files(max_age_hours: int = 24):
    """清理临时文件"""
    import glob
    import time
    
    temp_dir = tempfile.gettempdir()
    pattern = os.path.join(temp_dir, "*_template_*.xlsx")
    
    current_time = time.time()
    max_age_seconds = max_age_hours * 3600
    
    for filepath in glob.glob(pattern):
        try:
            if os.path.getctime(filepath) < current_time - max_age_seconds:
                os.remove(filepath)
                logger.info(f"已清理临时模板文件: {filepath}")
        except Exception as e:
            logger.warning(f"清理临时文件失败: {filepath}, 错误: {e}")