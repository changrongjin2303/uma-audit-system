import os
import uuid
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from fastapi import UploadFile, HTTPException
import xlrd
import openpyxl
from loguru import logger

from app.core.config import settings


class ExcelProcessor:
    """Excel文件处理器"""
    
    def detect_header_row(self, df_raw: pd.DataFrame, max_scan_rows: int = 10) -> Tuple[int, List[str]]:
        """
        智能检测表头行位置
        
        Args:
            df_raw: 原始DataFrame（包含所有行数据）
            max_scan_rows: 最大扫描行数
        
        Returns:
            (header_row_index, column_names): 表头行索引和列名列表
        """
        header_keywords = [
            # 中文关键字
            '材料', '名称', '规格', '型号', '单位', '价格', '单价', '数量', '金额', '备注',
            '编码', '序号', '项目', '工程', '费用', '成本', '预算', '造价', '合计', '小计',
            '品牌', '厂家', '供应商', '来源', '产地', '等级', '质量', '标准', '要求',
            # 英文关键字
            'material', 'name', 'specification', 'unit', 'price', 'quantity', 'amount',
            'code', 'no', 'item', 'project', 'cost', 'budget', 'total', 'remark'
        ]
        
        best_header_row = 0
        best_score = 0
        best_columns = []
        
        # 扫描前N行，寻找最可能的表头行
        scan_rows = min(max_scan_rows, len(df_raw))
        
        for row_idx in range(scan_rows):
            try:
                # 获取该行的所有值
                row_values = df_raw.iloc[row_idx].astype(str).tolist()
                
                # 过滤掉空值和无意义的值
                valid_values = [
                    str(val).strip() 
                    for val in row_values 
                    if pd.notna(val) and str(val).strip() and str(val).strip().lower() not in ['nan', 'none', '']
                ]
                
                if len(valid_values) < 2:  # 至少需要2个有效列名
                    continue
                
                # 计算该行作为表头的分数
                score = 0
                
                # 1. 关键字匹配分数
                for value in valid_values:
                    value_lower = value.lower()
                    for keyword in header_keywords:
                        if keyword in value_lower:
                            score += 10
                            break
                    
                    # 2. 中文字符比例分数（中文列名更可能是表头）
                    chinese_chars = sum(1 for char in value if '\u4e00' <= char <= '\u9fff')
                    if chinese_chars > 0:
                        score += min(chinese_chars * 2, 8)
                    
                    # 3. 长度适中分数（1-20个字符比较合理）
                    if 1 <= len(value) <= 20:
                        score += 3
                    
                    # 4. 避免纯数字（数字不太可能是列名）
                    try:
                        float(value)
                        score -= 5  # 纯数字减分
                    except ValueError:
                        score += 2  # 非纯数字加分
                
                # 5. 列数量分数（列数适中更可能是表头）
                if 3 <= len(valid_values) <= 20:
                    score += 5
                
                # 6. 连续性分数（连续的非空值更可能是表头）
                consecutive_non_empty = 0
                max_consecutive = 0
                for val in row_values:
                    if pd.notna(val) and str(val).strip():
                        consecutive_non_empty += 1
                        max_consecutive = max(max_consecutive, consecutive_non_empty)
                    else:
                        consecutive_non_empty = 0
                
                if max_consecutive >= 3:
                    score += max_consecutive * 2
                
                logger.info(f"行 {row_idx}: 分数={score}, 值={valid_values[:3]}...")
                
                if score > best_score:
                    best_score = score
                    best_header_row = row_idx
                    best_columns = [str(val).strip() for val in row_values if pd.notna(val)]
                    
            except Exception as e:
                logger.warning(f"检测第{row_idx}行时出错: {e}")
                continue
        
        logger.info(f"最佳表头行: 第{best_header_row}行, 分数={best_score}")
        return best_header_row, best_columns
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls', '.csv']
    
    def __init__(self):
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    async def validate_file(self, file: UploadFile) -> None:
        """验证上传的文件"""
        # 检查文件扩展名
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in self.SUPPORTED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件格式，支持的格式：{', '.join(self.SUPPORTED_EXTENSIONS)}"
            )
        
        # 检查文件大小
        content = await file.read()
        await file.seek(0)  # 重置文件指针
        
        if len(content) > settings.MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"文件大小超过限制（{settings.MAX_FILE_SIZE / 1024 / 1024:.1f}MB）"
            )
        
        # 检查文件是否为空
        if len(content) == 0:
            raise HTTPException(
                status_code=400,
                detail="上传的文件为空"
            )
    
    async def save_file(self, file: UploadFile) -> Tuple[str, str]:
        """保存上传的文件"""
        await self.validate_file(file)
        
        # 生成唯一文件名
        file_ext = Path(file.filename).suffix.lower()
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = self.upload_dir / unique_filename
        
        # 保存文件
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)
        
        logger.info(f"文件已保存: {file_path}")
        return str(file_path), unique_filename
    
    def read_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """读取Excel文件"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                # 读取CSV文件，尝试多种编码
                encodings = ['utf-8', 'gb2312', 'gbk', 'utf-8-sig']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise ValueError("无法解析CSV文件编码")
            
            elif file_ext == '.xlsx':
                # 读取Excel 2007+格式
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            
            elif file_ext == '.xls':
                # 读取Excel 97-2003格式
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
            
            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")
            
            return df
        
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            raise ValueError(f"读取Excel文件失败: {str(e)}")
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """获取Excel文件的工作表名称"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                return ['Sheet1']  # CSV文件只有一个工作表
            
            elif file_ext == '.xlsx':
                workbook = openpyxl.load_workbook(file_path)
                return workbook.sheetnames
            
            elif file_ext == '.xls':
                workbook = xlrd.open_workbook(file_path)
                return workbook.sheet_names()
            
            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")
        
        except Exception as e:
            logger.error(f"获取工作表名称失败: {e}")
            return []
    
    def analyze_columns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """分析Excel列结构"""
        analysis = {
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "column_info": [],
            "suggested_mapping": {}
        }
        
        # 标准化列名用于匹配
        standard_columns = {
            "material_name": ["材料名称", "名称", "品名", "材料", "项目名称"],
            "specification": ["规格", "规格型号", "型号", "规格参数", "技术参数"],
            "unit": ["单位", "计量单位", "单价单位"],
            "quantity": ["数量", "用量", "工程量"],
            "unit_price": ["单价", "综合单价", "材料单价", "不含税单价"],
            "total_price": ["合价", "总价", "金额", "不含税金额"],
            "category": ["分类", "类别", "专业", "分部分项"],
            "serial_number": ["序号", "编号", "项目编码", "清单编码"]
        }
        
        for idx, col in enumerate(df.columns):
            col_str = str(col).strip()
            
            # 分析列数据类型和统计信息
            # 获取示例值，过滤掉NaN和无穷大值
            sample_values = df[col].dropna().head(3).tolist()
            # 清理不能JSON序列化的值
            cleaned_sample_values = []
            for val in sample_values:
                if pd.isna(val) or (isinstance(val, (int, float)) and (val == float('inf') or val == float('-inf'))):
                    continue
                elif isinstance(val, float) and (val != val):  # 检查NaN
                    continue
                else:
                    cleaned_sample_values.append(val)
            
            col_info = {
                "index": idx,
                "name": col_str,
                "dtype": str(df[col].dtype),
                "null_count": int(df[col].isnull().sum()),  # 确保是普通int
                "unique_count": int(df[col].nunique()),     # 确保是普通int  
                "sample_values": cleaned_sample_values[:3]  # 最多3个示例值
            }
            
            analysis["column_info"].append(col_info)
            
            # 智能映射列名
            for standard_key, keywords in standard_columns.items():
                for keyword in keywords:
                    if keyword in col_str:
                        analysis["suggested_mapping"][standard_key] = col_str
                        break
        
        return analysis
    
    def parse_material_data(
        self, 
        df: pd.DataFrame, 
        column_mapping: Dict[str, str]
    ) -> List[Dict[str, Any]]:
        """解析材料数据"""
        materials = []
        
        # 验证必需的列是否存在
        required_columns = ["material_name", "unit"]
        for req_col in required_columns:
            if req_col not in column_mapping:
                raise ValueError(f"缺少必需的列映射: {req_col}")
        
        for index, row in df.iterrows():
            try:
                # 跳过空行或无效数据
                material_name = str(row.get(column_mapping.get("material_name", ""), "")).strip()
                if not material_name or material_name.lower() in ["nan", "none", ""]:
                    continue
                
                material = {
                    "row_number": index + 1,
                    "material_name": material_name,
                    "specification": self._safe_get_value(row, column_mapping.get("specification")),
                    "unit": self._safe_get_value(row, column_mapping.get("unit")),
                    "quantity": self._safe_get_numeric_value(row, column_mapping.get("quantity")),
                    "unit_price": self._safe_get_numeric_value(row, column_mapping.get("unit_price")),
                    "total_price": self._safe_get_numeric_value(row, column_mapping.get("total_price")),
                    "category": self._safe_get_value(row, column_mapping.get("category")),
                    "serial_number": self._safe_get_value(row, column_mapping.get("serial_number"))
                }
                
                # 计算总价（如果单价和数量都存在但总价不存在）
                if (material["unit_price"] and material["quantity"] and not material["total_price"]):
                    material["total_price"] = material["unit_price"] * material["quantity"]
                
                materials.append(material)
            
            except Exception as e:
                logger.warning(f"解析第{index + 1}行数据时出错: {e}")
                continue
        
        return materials
    
    def _safe_get_value(self, row: pd.Series, column_name: Optional[str]) -> Optional[str]:
        """安全获取字符串值"""
        if not column_name or column_name not in row:
            return None
        
        value = row[column_name]
        if pd.isna(value):
            return None
        
        return str(value).strip() if str(value).strip() else None
    
    def _safe_get_numeric_value(self, row: pd.Series, column_name: Optional[str]) -> Optional[float]:
        """安全获取数值"""
        if not column_name or column_name not in row:
            return None
        
        value = row[column_name]
        if pd.isna(value):
            return None
        
        try:
            # 尝试转换为数值
            if isinstance(value, (int, float)):
                return float(value)
            
            # 处理字符串形式的数值
            str_value = str(value).strip().replace(',', '').replace('¥', '').replace('￥', '')
            if str_value:
                return float(str_value)
        except (ValueError, TypeError):
            pass
        
        return None
    
    def validate_material_data(self, materials: List[Dict[str, Any]]) -> Dict[str, Any]:
        """验证材料数据质量"""
        validation_result = {
            "total_materials": len(materials),
            "valid_materials": 0,
            "warnings": [],
            "errors": []
        }
        
        for idx, material in enumerate(materials):
            # 检查必需字段
            if not material.get("material_name"):
                validation_result["errors"].append(f"第{material['row_number']}行: 材料名称为空")
                continue
            
            if not material.get("unit"):
                validation_result["warnings"].append(f"第{material['row_number']}行: 缺少计量单位")
            
            # 检查价格数据
            if material.get("unit_price") and material["unit_price"] <= 0:
                validation_result["warnings"].append(f"第{material['row_number']}行: 单价数据异常")
            
            if material.get("quantity") and material["quantity"] <= 0:
                validation_result["warnings"].append(f"第{material['row_number']}行: 数量数据异常")
            
            validation_result["valid_materials"] += 1
        
        return validation_result
    
    def get_full_data_for_import(self, file_content, filename: str, sheet_name: str = None, max_preview_rows: int = 100) -> Dict[str, Any]:
        """获取完整数据用于导入，只限制预览显示数量"""
        try:
            file_ext = Path(filename).suffix.lower()
            logger.info(f"获取预览数据: {filename}, 指定工作表: {sheet_name}")
            
            if file_ext == '.csv':
                try:
                    df = pd.read_csv(file_content, encoding='utf-8')
                except UnicodeDecodeError:
                    df = pd.read_csv(file_content, encoding='gbk')
                
                # 智能检测表头行
                df_raw = df.copy()
                header_row, detected_columns = self.detect_header_row(df_raw)
                
                # 重新读取，指定header行
                if hasattr(file_content, 'seek'):
                    file_content.seek(0)
                df = pd.read_csv(file_content, encoding='utf-8', header=header_row)
                
            else:
                # 处理Excel文件
                if file_ext == '.xlsx':
                    wb = openpyxl.load_workbook(file_content, read_only=True)
                    sheet_names = wb.sheetnames
                    target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else sheet_names[0]
                    wb.close()
                    
                    # 重新创建文件流
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    
                    # 先读取原始数据检测表头
                    df_raw = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl', header=None)
                    header_row, detected_columns = self.detect_header_row(df_raw)
                    
                    # 重新读取指定表头行
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    df = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl', header=header_row)
                    
                else:  # .xls
                    wb = xlrd.open_workbook(file_contents=file_content.read() if hasattr(file_content, 'read') else file_content)
                    sheet_names = [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
                    target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else 0
                    
                    # 先读取原始数据检测表头
                    df_raw = pd.read_excel(file_content, sheet_name=target_sheet, engine='xlrd', header=None)
                    header_row, detected_columns = self.detect_header_row(df_raw)
                    
                    # 重新读取指定表头行
                    df = pd.read_excel(file_content, sheet_name=target_sheet, engine='xlrd', header=header_row)
            
            columns = df.columns.tolist()
            
            # 生成全部数据用于导入
            full_data = []
            preview_data = []
            
            total_rows = len(df)
            preview_rows = min(total_rows, max_preview_rows)
            
            logger.info(f"开始处理数据: 总行数={total_rows}, 预览行数={preview_rows}")
            
            # 生成全部数据（用于实际导入）
            for idx in range(total_rows):
                row_data = df.iloc[idx]
                
                data_row = {
                    'row_index': idx,
                    'data': {}
                }
                
                # 生成列数据
                for i, col_name in enumerate(columns):
                    value = row_data[col_name]
                    
                    # 处理NaN值
                    if pd.isna(value):
                        value = ""
                    else:
                        value = str(value).strip()
                        
                    data_row['data'][col_name] = value
                    data_row[f'col_{i}'] = value  # 保持向后兼容
                
                full_data.append(data_row)
                
                # 前N行作为预览数据
                if idx < preview_rows:
                    preview_data.append(data_row)
            
            result = {
                "columns": columns,
                "previewData": preview_data,  # 用于前端预览显示
                "fullData": full_data,       # 用于实际导入
                "totalRows": total_rows,
                "previewRows": len(preview_data),
                "detectedHeaderRow": header_row if 'header_row' in locals() else 0,
                "currentSheet": target_sheet if 'target_sheet' in locals() else (sheet_name or "默认工作表")
            }
            
            logger.info(f"数据获取完成: 总行数{total_rows}, 完整数据{len(full_data)}, 预览行数{len(preview_data)}, 表头行{header_row if 'header_row' in locals() else 0}")
            return result
            
        except Exception as e:
            logger.error(f"获取预览数据失败: {e}")
            raise HTTPException(
                status_code=400,
                detail=f"获取预览数据失败: {str(e)}"
            )

    def analyze_file_structure(self, file_content, filename: str, sheet_name: str = None) -> Dict[str, Any]:
        """分析Excel文件结构"""
        try:
            file_ext = Path(filename).suffix.lower()
            logger.info(f"分析文件: {filename}, 指定工作表: {sheet_name}")
            
            if file_ext == '.csv':
                # 处理CSV文件
                try:
                    df = pd.read_csv(file_content, encoding='utf-8')
                except UnicodeDecodeError:
                    df = pd.read_csv(file_content, encoding='gbk')
                
                sheets = [{"name": "CSV数据", "rows": len(df), "columns": len(df.columns)}]
                columns = df.columns.tolist()
                sample_data = []
                
                # 获取前10行作为示例数据，并包含实际的列数据
                for idx, row in df.head(10).iterrows():
                    sample_row = {
                        'row_index': idx,
                        'data': {}
                    }
                    for i, col in enumerate(columns):
                        # 保持列名和数据的对应关系
                        sample_row['data'][col] = row[col]
                        sample_row[f'col_{i}'] = row[col]  # 保持向后兼容
                    sample_data.append(sample_row)
                
            else:
                # 处理Excel文件
                if file_ext == '.xlsx':
                    # 先获取所有工作表信息
                    wb = openpyxl.load_workbook(file_content, read_only=True)
                    sheet_names = wb.sheetnames
                    sheets = []
                    
                    for sheet_name_in_file in sheet_names:
                        ws = wb[sheet_name_in_file]
                        max_row = ws.max_row
                        max_col = ws.max_column
                        sheets.append({
                            "name": sheet_name_in_file,
                            "rows": max_row,
                            "columns": max_col
                        })
                    
                    wb.close()
                    
                    # 确定要读取的工作表
                    target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else sheet_names[0]
                    logger.info(f"读取工作表: {target_sheet}, 可用工作表: {sheet_names}")
                    
                    # 重新创建文件流并读取指定工作表
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    
                    # 先读取原始数据（不指定header）
                    df_raw = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl', header=None)
                    
                    # 智能检测表头行
                    header_row, detected_columns = self.detect_header_row(df_raw)
                    
                    # 重新读取文件指定表头行
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    
                    # 正确读取数据，设置列名
                    df = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl', header=header_row)
                    
                    # 如果检测到的表头行不是第0行，我们需要创建完整的样本数据
                    # 包括表头行本身，以便前端正确显示
                    if header_row > 0:
                        # 重新读取原始数据用于样本显示
                        if hasattr(file_content, 'seek'):
                            file_content.seek(0)
                        df_for_sample = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl', header=None)
                    else:
                        df_for_sample = df_raw
                    
                else:  # .xls
                    wb = xlrd.open_workbook(file_contents=file_content.read() if hasattr(file_content, 'read') else file_content)
                    sheets = []
                    
                    for sheet_idx in range(wb.nsheets):
                        sheet = wb.sheet_by_index(sheet_idx)
                        sheets.append({
                            "name": sheet.name,
                            "rows": sheet.nrows,
                            "columns": sheet.ncols
                        })
                    
                    # 使用指定的工作表或第一个工作表进行分析
                    sheet_names = [sheet.name for sheet in [wb.sheet_by_index(i) for i in range(wb.nsheets)]]
                    target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else 0
                    
                    # 先读取原始数据（不指定header）
                    df_raw = pd.read_excel(file_content, sheet_name=target_sheet, engine='xlrd', header=None)
                    
                    # 智能检测表头行
                    header_row, detected_columns = self.detect_header_row(df_raw)
                    
                    # 重新读取文件指定表头行
                    df = pd.read_excel(file_content, sheet_name=target_sheet, engine='xlrd', header=header_row)
                    
                    # 如果检测到的表头行不是第0行，重新读取用于样本显示
                    if header_row > 0:
                        df_for_sample = pd.read_excel(file_content, sheet_name=target_sheet, engine='xlrd', header=None)
                    else:
                        df_for_sample = df_raw
                
                columns = df.columns.tolist()
                sample_data = []
                
                # 使用正确的数据源生成样本数据
                sample_source = df_for_sample if 'df_for_sample' in locals() else df_raw if 'df_raw' in locals() else df
                
                # 确定样本数据的起始行（从表头行开始显示）
                start_row = header_row if 'header_row' in locals() and header_row is not None else 0
                end_row = min(start_row + 10, len(sample_source))
                
                # 生成样本数据，包含表头行和数据行
                for idx in range(start_row, end_row):
                    if idx >= len(sample_source):
                        break
                        
                    sample_row = {
                        'row_index': idx,
                        'data': {}
                    }
                    
                    # 获取该行的原始数据
                    if hasattr(sample_source, 'iloc'):
                        row_data = sample_source.iloc[idx]
                    else:
                        continue
                        
                    # 生成列数据
                    for i in range(len(columns)):
                        if i < len(row_data):
                            value = row_data.iloc[i] if hasattr(row_data, 'iloc') else (row_data[i] if i < len(row_data) else "")
                            # 处理NaN值
                            if pd.isna(value):
                                value = ""
                            else:
                                value = str(value).strip()
                        else:
                            value = ""
                            
                        # 使用实际的列名作为key
                        if i < len(columns):
                            sample_row['data'][columns[i]] = value
                        sample_row[f'col_{i}'] = value  # 保持向后兼容
                        
                    sample_data.append(sample_row)
                
                logger.info(f"样本数据生成完成: {len(sample_data)}行，从第{start_row}行开始")
            
            # 计算数据完整性
            total_cells = len(df) * len(df.columns)
            empty_cells = df.isnull().sum().sum()
            completeness = round((total_cells - empty_cells) / total_cells * 100, 2) if total_cells > 0 else 0
            
            # 当前实际读取的数据统计
            current_rows = len(df)
            current_columns = len(df.columns)
            
            logger.info(f"数据统计 - 工作表数量: {len(sheets)}, 当前工作表行数: {current_rows}, 列数: {current_columns}")
            
            result = {
                "sheets": sheets,
                "totalRows": current_rows,
                "totalColumns": current_columns,
                "completeness": completeness,
                "columns": columns,
                "sampleData": sample_data,
                "currentSheet": target_sheet if 'target_sheet' in locals() else (sheet_name or "默认工作表")
            }
            
            # 如果检测了表头行，添加相关信息
            if 'header_row' in locals() and header_row is not None:
                result["detectedHeaderRow"] = header_row
                result["headerDetectionApplied"] = True
                logger.info(f"智能表头检测: 第{header_row}行被识别为表头行")
            else:
                result["headerDetectionApplied"] = False
                
            return result
            
        except Exception as e:
            logger.error(f"分析文件结构失败: {e}")
            raise HTTPException(
                status_code=400,
                detail=f"文件结构分析失败: {str(e)}"
            )